import openpyxl as opx

def listaItens(name, colQtd, ixtable = 0, colName = 2, colUn = 3, colCod = 1):
    catologo = opx.load_workbook(filename=name)
    tabelaEscolha = catologo[catologo.sheetnames[ixtable]]
    lista = []
    for i in range(3, tabelaEscolha.max_row):
        line = {
                "codCliente": tabelaEscolha.cell(row=i, column=colCod).value,
                "nomeCliente": tabelaEscolha.cell(row=i, column=colName).value,
                "quantidade": tabelaEscolha.cell(row=i, column=colQtd).value,
                "unidade": tabelaEscolha.cell(row=i, column=colUn).value
            }
        if tabelaEscolha.cell(row=i, column=2).value != None:
            lista.append(line)
        
    return lista

def catBranet(name):
    catologo = opx.load_workbook(filename=name)
    tabelaEscolha = catologo[catologo.sheetnames[1]]
    lista = []
    for i in range(3, tabelaEscolha.max_row):
        line = {
                "codCliente": tabelaEscolha.cell(row=i, column=2).value,
                "nomeItem": tabelaEscolha.cell(row=i, column=3).value,
            }
        if tabelaEscolha.cell(row=i, column=3).value != None:
            lista.append(line)
            
    return lista
