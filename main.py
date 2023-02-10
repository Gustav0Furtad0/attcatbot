import scraping
import os
import openpyxl as opx
from colorama import Fore, Style, Back, init

init()

filePrefeitura = {}
fileBranet = ""
    
while True:

    while True:
        try:
            print("\n\n\n" + Back.WHITE + Fore.BLACK + "------- Arquivo prefeitura -------" + Style.RESET_ALL)
            filePrefeitura = {
                "fName": input("Digite o nome do arquivo: "),
                "ixTabela": int(input("Digite o índice da tabela: ")),
                "colunaCodCliente": int(input("Digite o índice da coluna que contenha o código do cliente: ")),
                "colunaNomeCliente": int(input("Digite o índice da coluna que contenha o nome do cliente: ")),
                "colunaUnidade": int(input("Digite o índice da coluna que contenha a unidade do item: ")),
                "colunaQuantidade": int(input("Digite o índice da coluna que contenha a quantidade: "))
            }

            print("\n\n\n" + Back.WHITE + Fore.BLACK + "------- Arquivo Branet -------" + Style.RESET_ALL)
            fileBranet = input("Digite o nome do arquivo: ")
            break
        
        except:
            print("Digite somente dados válidos...")

    if not (os.path.isfile(filePrefeitura["fName"]) or os.path.isfile(fileBranet)):
        print(Fore.RED + "Arquivo não encontrado, tente novamente..." + Style.RESET_ALL)
    else:
        try:
            catologo = opx.load_workbook(filename=filePrefeitura["fName"])
            cat = opx.load_workbook(filename=fileBranet)
            tabelaEscolha = catologo[catologo.sheetnames[filePrefeitura["ixTabela"]]]
            nomeUnidade = tabelaEscolha.cell(row=2, column=filePrefeitura["colunaQuantidade"]).value
            print("A unidade que deseja alterar o catálogo é " + Fore.GREEN + nomeUnidade.lstrip(" ") + Style.RESET_ALL + "? (y/n) ")
            tacerto = input()
            if tacerto == "y":
                break
            else:
                print(Fore.RED + 'Por favor, digite novamente os dados...' + Style.RESET_ALL)
        except:
            print(Fore.RED + "Arquivo não encontrado, tente novamente..." + Style.RESET_ALL)


navigator = scraping.ChromeDriver()

navigator.logaSistema('suportejf01', '123456789')

navigator.driver.get("https://juizdefora.branetlogistica.com.br/doms/processos/catalogo.xhtml")

input("Entre no catálogo que deseja fazer a modificações e pressione ENTER")

navigator.itens(fileBranet, filePrefeitura)